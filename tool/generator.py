"""
人材開発支援助成金（事業展開等リスキリング支援コース）書類自動生成エンジン
テンプレートのxlsxファイルをコピーし、ユーザー入力データで埋める
"""

import os
import shutil
import zipfile
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PLAN_DIR = os.path.join(BASE_DIR, "計画届（変更届）を提出する場合")
APP_DIR = os.path.join(BASE_DIR, "支給申請を行う場合")


def col_to_num(col_str):
    """A->1, B->2, ..., AA->27, etc."""
    num = 0
    for c in col_str:
        num = num * 26 + (ord(c.upper()) - ord('A') + 1)
    return num


def safe_write(ws, cell_ref, value):
    """セルに安全に書き込む（マージされたセルの場合はマージ解除なしで書込み）"""
    if value is None or value == "":
        return
    try:
        ws[cell_ref] = value
    except (AttributeError, ValueError):
        pass


def write_to_merged(ws, cell_ref, value):
    """マージされたセルの左上に書き込む"""
    if value is None or value == "":
        return
    for merged_range in ws.merged_cells.ranges:
        if cell_ref in merged_range:
            top_left = f"{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
            ws[top_left] = value
            return
    ws[cell_ref] = value


def set_checkbox(ws, cell_ref, checked=True):
    """チェックボックスを設定（☑ or □）"""
    write_to_merged(ws, cell_ref, "☑" if checked else "□")


def write_applicant_info(ws, data):
    """事業主情報を共通フォーマットで書き込む（法人・個人事業主対応）"""
    is_corporate = data.get("applicant_type", "corporate") == "corporate"

    write_to_merged(ws, "AG9", data.get("postal_code_1"))
    write_to_merged(ws, "AL9", data.get("postal_code_2"))
    write_to_merged(ws, "AF10", data.get("company_address"))

    if is_corporate:
        write_to_merged(ws, "AF12", data.get("company_name"))
        title = data.get("representative_title", "")
        name = data.get("representative_name", "")
        rep_str = (title + "　" + name).strip() if title else name
        write_to_merged(ws, "AF13", rep_str)
        if data.get("corporate_number"):
            write_to_merged(ws, "AF14", data.get("corporate_number"))
    else:
        # 個人事業主: 名称欄に屋号（あれば）、氏名欄に本人氏名
        if data.get("company_name"):
            write_to_merged(ws, "AF12", data.get("company_name"))
        write_to_merged(ws, "AF13", data.get("representative_name", ""))


def generate_form_1_1(data, output_path):
    """様式第1-1号 職業訓練実施計画届"""
    template = os.path.join(PLAN_DIR,
        "様式第1-1号人材開発支援助成金（事業展開等リスキリング支援コース）職業訓練実施計画届.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 提出日
    write_to_merged(ws, "AL5", data.get("submit_year"))
    write_to_merged(ws, "AR5", data.get("submit_month"))
    write_to_merged(ws, "AU5", data.get("submit_day"))

    # 労働局
    write_to_merged(ws, "B7", data.get("labor_bureau"))

    # 事業主情報
    write_applicant_info(ws, data)

    # 代理人（該当する場合）
    if data.get("has_agent"):
        write_to_merged(ws, "AG16", data.get("agent_postal_1"))
        write_to_merged(ws, "AL16", data.get("agent_postal_2"))
        write_to_merged(ws, "AF17", data.get("agent_address"))
        write_to_merged(ws, "AF19", data.get("agent_name_org"))
        write_to_merged(ws, "AF20", data.get("agent_name_person"))
        write_to_merged(ws, "AF21", data.get("agent_phone_1"))
        write_to_merged(ws, "AM21", data.get("agent_phone_2"))
        write_to_merged(ws, "AT21", data.get("agent_phone_3"))
        agent_type = data.get("agent_type", "代行")
        if agent_type == "代行":
            set_checkbox(ws, "Y20", True)
        else:
            set_checkbox(ws, "Y21", True)

    # 雇用保険適用事業所
    write_to_merged(ws, "K26", data.get("office_name"))
    write_to_merged(ws, "AN26", data.get("office_number_1"))
    write_to_merged(ws, "AS26", data.get("office_number_2"))
    # 1桁目 - AZ26付近
    safe_write(ws, "AZ26", data.get("office_number_3"))
    write_to_merged(ws, "K28", data.get("office_address"))
    write_to_merged(ws, "M27", data.get("office_postal_1"))
    write_to_merged(ws, "Q27", data.get("office_postal_2"))

    # 担当者
    write_to_merged(ws, "R29", data.get("contact_name"))
    write_to_merged(ws, "AM29", data.get("contact_dept"))
    write_to_merged(ws, "R30", data.get("contact_phone_1"))
    write_to_merged(ws, "W30", data.get("contact_phone_2"))
    write_to_merged(ws, "AB30", data.get("contact_phone_3"))
    write_to_merged(ws, "AM30", data.get("contact_email"))

    # 助成区分 - 事業展開等リスキリング支援コース
    subsidy_type = data.get("subsidy_type", "1")  # 1=事業展開, 2=DX, 3=グリーン
    is_subscription = data.get("is_subscription", False)

    if subsidy_type == "1":
        set_checkbox(ws, "K38", True)
    elif subsidy_type == "2":
        set_checkbox(ws, "Y38", True)
    elif subsidy_type == "3":
        set_checkbox(ws, "AM38", True)

    if is_subscription:
        set_checkbox(ws, "AF39", True)
    else:
        set_checkbox(ws, "K39", True)

    # 訓練コース名・受講者数
    write_to_merged(ws, "K41", data.get("course_name"))
    write_to_merged(ws, "AN41", data.get("num_trainees"))

    # 訓練の実施期間
    write_to_merged(ws, "N42", data.get("training_start_year"))
    write_to_merged(ws, "T42", data.get("training_start_month"))
    write_to_merged(ws, "Z42", data.get("training_start_day"))
    write_to_merged(ws, "AI42", data.get("training_end_year"))
    write_to_merged(ws, "AO42", data.get("training_end_month"))
    write_to_merged(ws, "AU42", data.get("training_end_day"))

    # 定額制サービスの契約期間
    if is_subscription:
        write_to_merged(ws, "N43", data.get("contract_start_year"))
        write_to_merged(ws, "T43", data.get("contract_start_month"))
        write_to_merged(ws, "Z43", data.get("contract_start_day"))
        write_to_merged(ws, "AI43", data.get("contract_end_year"))
        write_to_merged(ws, "AO43", data.get("contract_end_month"))
        write_to_merged(ws, "AU43", data.get("contract_end_day"))
        if data.get("auto_renewal"):
            set_checkbox(ws, "K44", True)

    # 資格試験
    if data.get("has_exam"):
        write_to_merged(ws, "O45", data.get("exam_name"))
        write_to_merged(ws, "AI45", data.get("exam_year"))
        write_to_merged(ws, "AO45", data.get("exam_month"))
        write_to_merged(ws, "AU45", data.get("exam_day"))

    # 訓練の実施場所
    write_to_merged(ws, "K46", data.get("training_location"))

    # 訓練の実施方法
    method = data.get("training_method", "1")  # 1=通学制, 2=同時双方向, 3=eラーニング, 4=通信制
    if method == "1":
        set_checkbox(ws, "K48", True)
    elif method == "2":
        set_checkbox(ws, "U48", True)
    elif method == "3":
        set_checkbox(ws, "AF48", True)
    elif method == "4":
        set_checkbox(ws, "AP48", True)

    # 訓練時間数
    if method in ("1", "2"):
        write_to_merged(ws, "R50", data.get("total_hours"))
        write_to_merged(ws, "Y50", data.get("total_minutes", "00"))
        write_to_merged(ws, "R51", data.get("offjt_hours"))
        write_to_merged(ws, "Y51", data.get("offjt_minutes", "00"))
    elif method in ("3", "4"):
        write_to_merged(ws, "R54", data.get("standard_hours"))
        write_to_merged(ws, "Y54", data.get("standard_minutes", "00"))

    # 第2面 - OFF-JT訓練種別
    offjt_type = data.get("offjt_type", "3")  # 1=部内講師, 2=部外講師, 3=事業外訓練
    if offjt_type == "1":
        set_checkbox(ws, "K59", True)
        write_to_merged(ws, "K60", data.get("instructor_name"))
    elif offjt_type == "2":
        set_checkbox(ws, "Y59", True)
        write_to_merged(ws, "K60", data.get("instructor_name"))
    elif offjt_type == "3":
        set_checkbox(ws, "AM59", True)

    # 教育訓練機関情報（事業外訓練の場合）
    if offjt_type in ("2", "3"):
        write_to_merged(ws, "R61", data.get("training_org_name"))
        write_to_merged(ws, "AM61", data.get("training_org_rep"))
        write_to_merged(ws, "R62", data.get("training_org_address"))

        # 契約経緯
        contract_reason = data.get("contract_reason", "2")
        if contract_reason == "1":
            set_checkbox(ws, "L65", True)
        elif contract_reason == "2":
            set_checkbox(ws, "L66", True)
        elif contract_reason == "3":
            set_checkbox(ws, "L67", True)
            write_to_merged(ws, "S67", data.get("contract_reason_other", ""))

        # 負担軽減チェック
        set_checkbox(ws, "L71", True)

    wb.save(output_path)
    wb.close()


def generate_form_1_3(data, output_path):
    """様式第1-3号 事業展開等実施計画"""
    template = os.path.join(PLAN_DIR,
        "様式第1-3号人材開発支援助成金（事業展開等リスキリング支援コース）事業展開等実施計画.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    subsidy_type = data.get("subsidy_type", "1")

    # 事業展開等の種類チェック
    if subsidy_type == "1":
        set_checkbox(ws, "P8", True)
    else:
        set_checkbox(ws, "P9", True)

    # 事業展開の実施（予定）時期
    write_to_merged(ws, "B13", data.get("expansion_year"))
    write_to_merged(ws, "E13", data.get("expansion_month"))

    # 事業展開の内容 or DX化の内容
    if subsidy_type == "1":
        write_to_merged(ws, "A19", data.get("expansion_content"))
    else:
        write_to_merged(ws, "A29", data.get("dx_content"))

    # 証明日付
    write_to_merged(ws, "H37", data.get("cert_year"))
    write_to_merged(ws, "L37", data.get("cert_month"))
    write_to_merged(ws, "O37", data.get("cert_day"))

    # 代表者
    write_to_merged(ws, "K40", data.get("representative_title", "") + "　" + data.get("company_name", ""))
    write_to_merged(ws, "K41", data.get("representative_name"))

    wb.save(output_path)
    wb.close()


def generate_form_3_1(data, output_path):
    """様式第3-1号 対象労働者一覧"""
    template = os.path.join(PLAN_DIR,
        "様式第3-1号人材開発支援助成金（事業展開等リスキリング支援コース）対象労働者一覧.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # ページ番号
    safe_write(ws, "V3", "1")
    safe_write(ws, "AA3", "1")

    # 事業所名、訓練コース名
    write_to_merged(ws, "K8", data.get("office_name"))  # Row 8 has input area
    write_to_merged(ws, "K9", data.get("course_name"))

    # 労働者一覧
    workers = data.get("workers", [])
    base_row = 13  # 1人目は行13から
    row_step = 2   # 2行ずつ（正規/有期の選択肢行があるため）

    for i, w in enumerate(workers):
        row = base_row + (i * row_step)
        if row > 107:  # フォームの限界
            break
        safe_write(ws, f"B{row}", w.get("name"))
        safe_write(ws, f"D{row}", w.get("insurance_1"))
        safe_write(ws, f"E{row}", w.get("insurance_1"))
        safe_write(ws, f"H{row}", w.get("insurance_2"))
        # 雇用形態チェック
        emp_type = w.get("employment_type", "regular")
        if emp_type == "regular":
            set_checkbox(ws, f"J{row}", True)
        else:
            set_checkbox(ws, f"J{row+1}", True)

    wb.save(output_path)
    wb.close()


def generate_form_3_2(data, output_path):
    """様式第3-2号 定額制サービスによる訓練に関する対象労働者一覧"""
    template = os.path.join(PLAN_DIR, "様式第3-2号.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # ページ番号
    safe_write(ws, "G3", "1")
    safe_write(ws, "I3", "1")

    # 証明日付
    safe_write(ws, "E8", data.get("submit_year"))
    safe_write(ws, "G8", data.get("submit_month"))
    safe_write(ws, "I8", data.get("submit_day"))

    # 事業所名、訓練コース名
    write_to_merged(ws, "B11", data.get("office_name"))
    write_to_merged(ws, "B12", data.get("course_name"))

    # 労働者一覧
    workers = data.get("workers", [])
    for i, w in enumerate(workers):
        row = 16 + i
        if row > 75:
            break
        safe_write(ws, f"B{row}", w.get("name"))
        emp_type = w.get("employment_type", "regular")
        if emp_type == "regular":
            set_checkbox(ws, f"C{row}", True)
        else:
            set_checkbox(ws, f"G{row}", True)

    wb.save(output_path)
    wb.close()


def generate_form_11(data, output_path):
    """様式第11号 事前確認書"""
    template = os.path.join(PLAN_DIR,
        "様式第11号人材開発支援助成金（事業展開等リスキリング支援コース）事前確認書.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 提出日
    safe_write(ws, "E12", data.get("submit_year"))
    safe_write(ws, "H12", data.get("submit_month"))
    safe_write(ws, "K12", data.get("submit_day"))

    # 事業主情報
    write_to_merged(ws, "S13", data.get("postal_code_1"))
    write_to_merged(ws, "W13", data.get("postal_code_2"))
    write_to_merged(ws, "O14", data.get("company_address"))
    write_to_merged(ws, "O16", data.get("company_name"))
    write_to_merged(ws, "O18", data.get("representative_title", "") + "　" + data.get("representative_name", ""))
    write_to_merged(ws, "O19", data.get("contact_phone_1"))
    write_to_merged(ws, "V19", data.get("contact_phone_2"))
    write_to_merged(ws, "Z19", data.get("contact_phone_3"))

    # 労働局
    safe_write(ws, "C30", data.get("labor_bureau"))

    # 全チェック項目にチェックを入れる（事前確認なので全て☑）
    # 注意事項の確認チェック欄は元のフォームの構造に依存

    wb.save(output_path)
    wb.close()


def generate_form_4_2(data, output_path):
    """様式第4-2号 支給申請書"""
    template = os.path.join(APP_DIR,
        "様式第4-2号人材開発支援助成金（事業展開等リスキリング支援コース）支給申請書.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 申請日
    write_to_merged(ws, "AL5", data.get("app_year"))
    write_to_merged(ws, "AR5", data.get("app_month"))
    write_to_merged(ws, "AU5", data.get("app_day"))

    # 労働局
    write_to_merged(ws, "B7", data.get("labor_bureau"))

    # 事業主情報（様式1-1号と同じ構造）
    write_applicant_info(ws, data)

    # 代理人
    if data.get("has_agent"):
        write_to_merged(ws, "AG16", data.get("agent_postal_1"))
        write_to_merged(ws, "AL16", data.get("agent_postal_2"))
        write_to_merged(ws, "AF17", data.get("agent_address"))
        write_to_merged(ws, "AF19", data.get("agent_name_org"))
        write_to_merged(ws, "AF20", data.get("agent_name_person"))
        write_to_merged(ws, "AF21", data.get("agent_phone_1"))
        write_to_merged(ws, "AM21", data.get("agent_phone_2"))
        write_to_merged(ws, "AT21", data.get("agent_phone_3"))

    # 計画届の受付番号
    write_to_merged(ws, "K25", data.get("plan_receipt_number"))

    # 主たる事業
    write_to_merged(ws, "K26", data.get("main_business"))

    # 常時雇用する労働者数
    write_to_merged(ws, "K27", data.get("total_employees"))

    # 雇用保険適用事業所
    write_to_merged(ws, "K28", data.get("office_name"))
    write_to_merged(ws, "AN28", data.get("office_number_1"))
    write_to_merged(ws, "AS28", data.get("office_number_2"))
    safe_write(ws, "AZ28", data.get("office_number_3"))

    wb.save(output_path)
    wb.close()


def generate_form_5(data, output_path):
    """様式第5号 賃金助成の内訳"""
    template = os.path.join(APP_DIR,
        "様式第5号人材開発支援助成金（事業展開等リスキリング支援コース）賃金助成の内訳.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 受付番号・事業所名
    write_to_merged(ws, "K7", data.get("plan_receipt_number"))
    write_to_merged(ws, "BA7", data.get("office_name"))

    # 賃金助成対象時間数
    write_to_merged(ws, "D11", data.get("wage_subsidy_hours"))
    write_to_merged(ws, "H11", data.get("wage_subsidy_minutes", "00"))

    # 賃金助成の単価
    is_sme = data.get("is_sme", True)
    unit_price = 1000 if is_sme else 500
    write_to_merged(ws, "O10", unit_price)

    # 対象労働者一覧
    workers = data.get("workers", [])
    base_row = 21
    for i, w in enumerate(workers):
        row = base_row + i
        if row > 120:
            break
        safe_write(ws, f"B{row}", w.get("name"))
        safe_write(ws, f"C{row}", w.get("name_kana"))
        safe_write(ws, f"J{row}", w.get("insurance_number"))

    wb.save(output_path)
    wb.close()


def generate_form_6_2(data, output_path):
    """様式第6-2号 経費助成の内訳"""
    template = os.path.join(APP_DIR,
        "様式第6-2号人材開発支援助成金（事業展開等リスキリング支援コース）経費助成の内訳.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 受付番号・事業所名
    write_to_merged(ws, "K7", data.get("plan_receipt_number"))
    write_to_merged(ws, "AD7", data.get("office_name"))

    # 経費情報は訓練タイプにより異なる
    offjt_type = data.get("offjt_type", "3")
    if offjt_type in ("1", "2"):  # 事業内訓練
        safe_write(ws, "E12", data.get("instructor_fee", 0))
        safe_write(ws, "M12", data.get("travel_fee", 0))
        safe_write(ws, "U12", data.get("facility_fee", 0))
        safe_write(ws, "AC12", data.get("material_fee", 0))
        safe_write(ws, "AK12", data.get("development_fee", 0))

    wb.save(output_path)
    wb.close()


def generate_form_6_3(data, output_path):
    """様式第6-3号 定額制サービスによる訓練に関する経費助成の内訳"""
    template = os.path.join(APP_DIR,
        "様式第6-3号人材開発支援助成金（事業展開等リスキリング支援コース） 定額制サービスによる訓練に関する経費助成の内訳.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 助成区分 - 事業展開等リスキリング支援コースをチェック
    set_checkbox(ws, "AJ5", True)

    # 受付番号
    write_to_merged(ws, "L6", data.get("plan_receipt_number"))

    # 訓練コース名
    write_to_merged(ws, "AE6", data.get("course_name"))

    # 助成対象労働者数
    write_to_merged(ws, "L7", data.get("num_trainees"))

    # 契約者数
    write_to_merged(ws, "AE7", data.get("total_subscribers"))

    # 訓練の実施期間
    write_to_merged(ws, "M8", data.get("training_start_year"))
    write_to_merged(ws, "R8", data.get("training_start_month"))
    write_to_merged(ws, "W8", data.get("training_start_day"))
    write_to_merged(ws, "AE8", data.get("training_end_year"))
    write_to_merged(ws, "AJ8", data.get("training_end_month"))
    write_to_merged(ws, "AO8", data.get("training_end_day"))

    wb.save(output_path)
    wb.close()


def generate_form_8_1(data, output_path):
    """様式第8-1号 OFF-JT実施状況報告書"""
    template = os.path.join(APP_DIR,
        "様式第8-1号人材開発支援助成金（事業展開等リスキリング支援コース）OFF-JT実施状況報告書保護解除.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 受付番号・訓練コース名
    write_to_merged(ws, "K6", data.get("plan_receipt_number"))
    write_to_merged(ws, "AE6", data.get("course_name"))

    # OFF-JT種別
    offjt_type = data.get("offjt_type", "3")
    if offjt_type in ("1", "2"):
        set_checkbox(ws, "K7", True)
    else:
        set_checkbox(ws, "S7", True)

    # 教育訓練機関名
    if offjt_type == "3":
        write_to_merged(ws, "AE7", data.get("training_org_name"))

    wb.save(output_path)
    wb.close()


def generate_form_8_3(data, output_path):
    """様式第8-3号 eラーニング訓練実施結果報告書"""
    template = os.path.join(APP_DIR,
        "様式第8-3号人材開発支援助成金（事業展開等リスキリング支援コース）eラーニング訓練実施結果報告書.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 事業所名
    write_to_merged(ws, "K5", data.get("office_name"))

    # 訓練期間
    write_to_merged(ws, "K6", data.get("training_start_year"))
    write_to_merged(ws, "N6", data.get("training_start_month"))
    write_to_merged(ws, "Q6", data.get("training_start_day"))
    write_to_merged(ws, "Y6", data.get("training_end_year"))
    write_to_merged(ws, "AB6", data.get("training_end_month"))
    write_to_merged(ws, "AE6", data.get("training_end_day"))

    wb.save(output_path)
    wb.close()


def generate_form_12(data, output_path):
    """様式第12号 支給申請承諾書（訓練実施者）"""
    template = os.path.join(APP_DIR,
        "様式第12号人材開発支援助成金（事業展開等リスキリング支援コース）支給申請承諾書（訓練実施者）.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 労働局
    safe_write(ws, "B24", data.get("labor_bureau"))

    # 確認日
    write_to_merged(ws, "R24", data.get("app_year"))
    write_to_merged(ws, "U24", data.get("app_month"))
    write_to_merged(ws, "X24", data.get("app_day"))

    # 教育訓練機関情報
    write_to_merged(ws, "E26", data.get("training_org_address"))
    write_to_merged(ws, "E28", data.get("training_org_name"))
    write_to_merged(ws, "E30", data.get("training_org_rep"))
    write_to_merged(ws, "E32", data.get("training_org_corp_number"))

    # 訓練情報
    write_to_merged(ws, "K39", data.get("plan_receipt_number"))
    write_to_merged(ws, "Q39", data.get("course_name"))

    wb.save(output_path)
    wb.close()


def generate_form_13(data, output_path):
    """様式第13号 事業所確認票"""
    template = os.path.join(APP_DIR,
        "様式第13号人材開発支援助成金（事業展開等リスキリング支援コース）事業所確認票.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 提出日
    write_to_merged(ws, "N3", data.get("app_year"))
    write_to_merged(ws, "P3", data.get("app_month"))
    write_to_merged(ws, "R3", data.get("app_day"))

    # 労働局
    safe_write(ws, "A4", data.get("labor_bureau"))

    # 事業主
    write_to_merged(ws, "L8", data.get("company_name"))
    write_to_merged(ws, "L10", data.get("company_address"))

    # 事業所
    write_to_merged(ws, "B15", data.get("office_name"))
    write_to_merged(ws, "H16", data.get("office_number_1"))
    write_to_merged(ws, "L16", data.get("office_number_2"))
    safe_write(ws, "P16", data.get("office_number_3"))
    write_to_merged(ws, "Q15", data.get("total_employees"))

    wb.save(output_path)
    wb.close()


def generate_form_10(data, output_path):
    """様式第10号 OFF-JT講師要件確認書"""
    template = os.path.join(PLAN_DIR,
        "様式第10号人材開発支援助成金（事業展開等リスキリング支援コース）OFF-JT講師要件確認書.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 日付
    safe_write(ws, "AA5", data.get("submit_year"))
    safe_write(ws, "AD5", data.get("submit_month"))
    safe_write(ws, "AG5", data.get("submit_day"))

    # 講師情報
    offjt_type = data.get("offjt_type", "3")
    if offjt_type == "1":  # 部内講師
        write_to_merged(ws, "I8", data.get("instructor_name"))
        write_to_merged(ws, "I9", data.get("instructor_dept"))
        write_to_merged(ws, "I10", data.get("instructor_title"))
        write_to_merged(ws, "I11", data.get("instructor_duties"))
    elif offjt_type == "2":  # 部外講師
        write_to_merged(ws, "AE8", data.get("instructor_name"))

    wb.save(output_path)
    wb.close()


def generate_form_2_1(data, output_path):
    """様式第2-1号 職業訓練実施計画変更届"""
    template = os.path.join(PLAN_DIR,
        "様式第2-1号人材開発支援助成金（事業展開等リスキリング支援コース）職業訓練実施計画変更届.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 提出日
    write_to_merged(ws, "AQ5", data.get("submit_year"))
    write_to_merged(ws, "AT5", data.get("submit_month"))
    write_to_merged(ws, "AW5", data.get("submit_day"))

    # 労働局
    write_to_merged(ws, "B7", data.get("labor_bureau"))

    # 事業主情報
    write_to_merged(ws, "AF9", data.get("postal_code_1"))
    write_to_merged(ws, "AK9", data.get("postal_code_2"))
    write_to_merged(ws, "AC10", data.get("company_address"))
    is_corporate = data.get("applicant_type", "corporate") == "corporate"
    if is_corporate:
        write_to_merged(ws, "AC12", data.get("company_name"))
        title = data.get("representative_title", "")
        name = data.get("representative_name", "")
        rep_str = (title + "　" + name).strip() if title else name
        write_to_merged(ws, "AC13", rep_str)
    else:
        if data.get("company_name"):
            write_to_merged(ws, "AC12", data.get("company_name"))
        write_to_merged(ws, "AC13", data.get("representative_name", ""))

    # 受付番号
    write_to_merged(ws, "K25", data.get("plan_receipt_number"))

    # 事業所名・番号
    write_to_merged(ws, "K26", data.get("office_name"))
    write_to_merged(ws, "AN26", data.get("office_number_1"))
    write_to_merged(ws, "AR26", data.get("office_number_2"))
    safe_write(ws, "AZ26", data.get("office_number_3"))

    # 担当者
    write_to_merged(ws, "K27", data.get("contact_name"))
    write_to_merged(ws, "AF27", data.get("contact_dept"))
    write_to_merged(ws, "K28", data.get("contact_phone_1"))
    write_to_merged(ws, "V28", data.get("contact_phone_2"))
    write_to_merged(ws, "AA28", data.get("contact_phone_3"))
    write_to_merged(ws, "AF28", data.get("contact_email"))

    # 助成区分 - 事業展開等リスキリング支援コース
    subsidy_type = data.get("subsidy_type", "1")
    is_subscription = data.get("is_subscription", False)
    if subsidy_type == "1":
        set_checkbox(ws, "M36", True)
    elif subsidy_type == "2":
        set_checkbox(ws, "AA36", True)
    elif subsidy_type == "3":
        set_checkbox(ws, "AO36", True)
    if is_subscription:
        set_checkbox(ws, "AH37", True)
    else:
        set_checkbox(ws, "M37", True)

    # コース名・受講者数
    write_to_merged(ws, "K40", data.get("course_name"))
    write_to_merged(ws, "AG40", data.get("num_trainees"))

    # 訓練期間
    write_to_merged(ws, "N42", data.get("training_start_year"))
    write_to_merged(ws, "T42", data.get("training_start_month"))
    write_to_merged(ws, "Z42", data.get("training_start_day"))
    write_to_merged(ws, "AI42", data.get("training_end_year"))
    write_to_merged(ws, "AO42", data.get("training_end_month"))
    write_to_merged(ws, "AU42", data.get("training_end_day"))

    # 実施場所
    write_to_merged(ws, "K45", data.get("training_location"))

    # 実施方法
    method = data.get("training_method", "1")
    if method == "1":
        set_checkbox(ws, "M47", True)
    elif method == "2":
        set_checkbox(ws, "W47", True)
    elif method == "3":
        set_checkbox(ws, "AH47", True)
    elif method == "4":
        set_checkbox(ws, "AR47", True)

    # 変更理由
    write_to_merged(ws, "B63", data.get("change_reason", ""))

    wb.save(output_path)
    wb.close()


def generate_form_14_1(data, output_path):
    """様式第14-1号 定額制サービスによる訓練に関する事業所確認票"""
    template = os.path.join(PLAN_DIR,
        "様式第14-1号人材開発支援助成金（事業展開等リスキリング支援コース） 定額制サービスによる訓練に関する事業所確認票.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 提出日
    safe_write(ws, "R3", data.get("submit_year"))
    safe_write(ws, "U3", data.get("submit_month"))
    safe_write(ws, "W3", data.get("submit_day"))

    # 労働局
    safe_write(ws, "B4", data.get("labor_bureau"))

    # 事業主
    write_to_merged(ws, "L8", data.get("company_name"))
    write_to_merged(ws, "L10", data.get("company_address"))

    # 訓練コース名
    write_to_merged(ws, "F12", data.get("course_name"))

    # 申請事業所
    write_to_merged(ws, "B16", data.get("office_name"))
    write_to_merged(ws, "G16", data.get("office_number_1"))
    write_to_merged(ws, "L16", data.get("office_number_2"))
    safe_write(ws, "S16", data.get("office_number_3"))
    write_to_merged(ws, "T16", data.get("num_trainees"))

    # 他事業所から申請しないチェック
    set_checkbox(ws, "B46", True)

    wb.save(output_path)
    wb.close()


def generate_form_14_2(data, output_path):
    """様式第14-2号 本社一括申請に関する事業所確認票"""
    template = os.path.join(PLAN_DIR,
        "様式第14-2号人材開発支援助成金（事業展開等リスキリング支援コース） 本社一括申請に関する事業所確認票.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 提出日
    safe_write(ws, "T3", data.get("submit_year"))
    safe_write(ws, "W3", data.get("submit_month"))
    safe_write(ws, "Y3", data.get("submit_day"))

    # 労働局
    safe_write(ws, "B5", data.get("labor_bureau"))

    # 事業主
    write_to_merged(ws, "M9", data.get("company_name"))
    write_to_merged(ws, "M10", data.get("company_address"))

    # 訓練コース名
    write_to_merged(ws, "F12", data.get("course_name"))

    # 本社事業所
    write_to_merged(ws, "B17", data.get("office_name"))
    write_to_merged(ws, "G17", data.get("office_number_1"))
    write_to_merged(ws, "L17", data.get("office_number_2"))
    safe_write(ws, "S17", data.get("office_number_3"))

    # 一括申請チェック
    set_checkbox(ws, "B45", True)

    wb.save(output_path)
    wb.close()


def generate_form_7(data, output_path):
    """様式第7号 自発的職業能力開発に関する申立書"""
    template = os.path.join(APP_DIR,
        "様式第7号人材開発支援助成金（事業展開等リスキリング支援コース）自発的職業能力開発に関する申立書.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # コース名・機関名
    write_to_merged(ws, "I9", data.get("course_name"))
    write_to_merged(ws, "I10", data.get("training_org_name"))

    # 受講料
    write_to_merged(ws, "I11", data.get("total_training_fee"))
    write_to_merged(ws, "I12", data.get("employer_fee_share"))
    write_to_merged(ws, "I13", data.get("worker_fee_share", "0"))

    # 日付
    safe_write(ws, "A21", data.get("app_year"))
    safe_write(ws, "E21", data.get("app_month"))
    safe_write(ws, "H21", data.get("app_day"))

    # 労働局
    safe_write(ws, "B23", data.get("labor_bureau"))

    wb.save(output_path)
    wb.close()


def generate_form_8_4(data, output_path):
    """様式第8-4号 通信制訓練実施結果報告書"""
    template = os.path.join(APP_DIR,
        "様式第8-4号人材開発支援助成金（事業展開等リスキリング支援コース）通信制訓練実施結果報告書.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 事業所名
    write_to_merged(ws, "C6", data.get("office_name"))

    # 訓練コース名
    write_to_merged(ws, "C7", data.get("course_name"))

    # 訓練期間
    safe_write(ws, "J8", data.get("training_start_year"))
    safe_write(ws, "N8", data.get("training_start_month"))
    safe_write(ws, "Q8", data.get("training_start_day"))
    safe_write(ws, "X8", data.get("training_end_year"))
    safe_write(ws, "AB8", data.get("training_end_month"))
    safe_write(ws, "AE8", data.get("training_end_day"))

    wb.save(output_path)
    wb.close()


def generate_form_8_5(data, output_path):
    """様式第8-5号 定額制サービスによる訓練実施結果報告書"""
    template = os.path.join(APP_DIR,
        "様式第8-5号人材開発支援助成金（事業展開等リスキリング支援コース） 定額制サービスによる訓練実施結果報告書.xlsx")
    shutil.copy2(template, output_path)
    wb = load_workbook(output_path)
    ws = wb[wb.sheetnames[0]]

    # 事業所名
    write_to_merged(ws, "A5", data.get("office_name"))

    wb.save(output_path)
    wb.close()


def generate_all_documents(data):
    """全書類を生成してZIPにまとめる"""
    # Vercel環境では/tmpに出力、ローカルではtool/output
    if os.environ.get("VERCEL"):
        output_dir = "/tmp/jinzai_output"
    else:
        output_dir = os.path.join(BASE_DIR, "tool", "output")
    # 出力ディレクトリをクリア
    if os.path.exists(output_dir):
        shutil.rmtree(output_dir)
    os.makedirs(output_dir, exist_ok=True)

    plan_dir = os.path.join(output_dir, "01_計画届")
    app_dir_out = os.path.join(output_dir, "02_支給申請")
    os.makedirs(plan_dir, exist_ok=True)
    os.makedirs(app_dir_out, exist_ok=True)

    generated_files = []
    is_subscription = data.get("is_subscription", False)
    offjt_type = data.get("offjt_type", "3")
    training_method = data.get("training_method", "1")

    # === 計画届 ===
    # 様式第1-1号（必須）
    path = os.path.join(plan_dir, "様式第1-1号_職業訓練実施計画届.xlsx")
    generate_form_1_1(data, path)
    generated_files.append(path)

    # 様式第1-3号（必須）
    path = os.path.join(plan_dir, "様式第1-3号_事業展開等実施計画.xlsx")
    generate_form_1_3(data, path)
    generated_files.append(path)

    # 様式第3-1号 or 3-2号（必須）
    if is_subscription:
        path = os.path.join(plan_dir, "様式第3-2号_定額制対象労働者一覧.xlsx")
        generate_form_3_2(data, path)
    else:
        path = os.path.join(plan_dir, "様式第3-1号_対象労働者一覧.xlsx")
        generate_form_3_1(data, path)
    generated_files.append(path)

    # 様式第11号（必須）
    path = os.path.join(plan_dir, "様式第11号_事前確認書.xlsx")
    generate_form_11(data, path)
    generated_files.append(path)

    # 様式第10号（事業内訓練の場合）
    if offjt_type in ("1", "2"):
        path = os.path.join(plan_dir, "様式第10号_OFF-JT講師要件確認書.xlsx")
        generate_form_10(data, path)
        generated_files.append(path)

    # 様式第14-1号（定額制サービスの場合）
    if is_subscription:
        path = os.path.join(plan_dir, "様式第14-1号_定額制サービス事業所確認票.xlsx")
        generate_form_14_1(data, path)
        generated_files.append(path)

    # 様式第14-2号（本社一括申請の場合）
    if data.get("is_batch_application", False):
        path = os.path.join(plan_dir, "様式第14-2号_本社一括申請事業所確認票.xlsx")
        generate_form_14_2(data, path)
        generated_files.append(path)

    # === 支給申請 ===
    # 様式第4-2号（必須）
    path = os.path.join(app_dir_out, "様式第4-2号_支給申請書.xlsx")
    generate_form_4_2(data, path)
    generated_files.append(path)

    # 様式第5号 賃金助成の内訳（通学制/同時双方向の場合）
    if training_method in ("1", "2"):
        path = os.path.join(app_dir_out, "様式第5号_賃金助成の内訳.xlsx")
        generate_form_5(data, path)
        generated_files.append(path)

    # 様式第6-2号 or 6-3号 経費助成
    if is_subscription:
        path = os.path.join(app_dir_out, "様式第6-3号_定額制経費助成の内訳.xlsx")
        generate_form_6_3(data, path)
    else:
        path = os.path.join(app_dir_out, "様式第6-2号_経費助成の内訳.xlsx")
        generate_form_6_2(data, path)
    generated_files.append(path)

    # 様式第7号（自発的職業能力開発の場合）
    if data.get("is_voluntary", False):
        path = os.path.join(app_dir_out, "様式第7号_自発的職業能力開発申立書.xlsx")
        generate_form_7(data, path)
        generated_files.append(path)

    # 様式第8系 実施状況報告書
    if training_method in ("1", "2"):
        path = os.path.join(app_dir_out, "様式第8-1号_OFF-JT実施状況報告書.xlsx")
        generate_form_8_1(data, path)
        generated_files.append(path)
    elif training_method == "3":
        path = os.path.join(app_dir_out, "様式第8-3号_eラーニング訓練実施結果報告書.xlsx")
        generate_form_8_3(data, path)
        generated_files.append(path)
    elif training_method == "4":
        path = os.path.join(app_dir_out, "様式第8-4号_通信制訓練実施結果報告書.xlsx")
        generate_form_8_4(data, path)
        generated_files.append(path)

    # 様式第8-5号（定額制サービスの場合）
    if is_subscription:
        path = os.path.join(app_dir_out, "様式第8-5号_定額制訓練実施結果報告書.xlsx")
        generate_form_8_5(data, path)
        generated_files.append(path)

    # 様式第12号 支給申請承諾書（事業外訓練の場合）
    if offjt_type == "3":
        path = os.path.join(app_dir_out, "様式第12号_支給申請承諾書.xlsx")
        generate_form_12(data, path)
        generated_files.append(path)

    # 様式第13号 事業所確認票
    if data.get("is_sme", True):
        path = os.path.join(app_dir_out, "様式第13号_事業所確認票.xlsx")
        generate_form_13(data, path)
        generated_files.append(path)

    # ZIPにまとめる
    zip_path = os.path.join(output_dir, "人材開発支援助成金_申請書類一式.zip")
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fp in generated_files:
            arcname = os.path.relpath(fp, output_dir)
            zf.write(fp, arcname)

    return zip_path, generated_files
